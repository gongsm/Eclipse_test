#-*- coding: UTF-8 -*-
'''
合并excel的两个sheet
'''
import sys
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, Font, Side, Border, PatternFill, Style
from openpyxl.reader.excel import load_workbook
from bunch import Bunch

signalsInputColumns = (
    # header,                         value,                         len
      ("Skip",                        "Skip",                         5),                                                             
      ("RpName",                      "RpName",                      40),
      ("Pubref",                      "Pubref",                      40),
      ("TxLru",                       "TxLru",                       30),
      ("Message",                     "Message",                     40),
      ("DataSet",                     "DataSet",                     40),
      ("DpName",                      "DpName",                      40),
      ("ValidityCriteria",            "ValidityCriteria",            15),                                                                         
      ("FsbOffset",                   "FsbOffset",                   10),
      ("DSOffset",                    "DSOffset",                    10),
      ("DSSize",                      "DSSize",                      10),
      ("ParameterType",               "ParameterType",               10),
      ("BitOffsetWithinDS",           "BitOffsetWithinDS",           10),
      ("ParameterSize",               "ParameterSize",               10),
      ("LsbRes",                      "LsbRes",                      10),
      ("PublisherFunctionalMinRange", "PublisherFunctionalMinRange", 35),
      ("PublisherFunctionalMaxRange", "PublisherFunctionalMaxRange", 35),
      ("Comment",                     "Comment",                     40),
      ("Status",                      "Status",                      40),      
      ("ChangeHistory",               "ChangeHistory",               40),      
      ("ICDFix",                      "ICDFix",                      40),      
      ("MessageRef",                  "MessageRef",                  40),
)

def readExcelFile(wb, sheets):
    resultlst = []
    for title in sheets:
        ws = wb.get_sheet_by_name(title)
        if not ws:
            return None

        rowidx = 0
        reclst = []
        for row in ws.iter_rows():
            if rowidx == 0:
                # read header
                header = [str(cell.value).strip() for cell in row]
            else:
                # get a line, skip if first column starts with #
                rec = Bunch()
                colidx = 0
                for cell in row:
                    rec[header[colidx]] = cell.value
                    colidx += 1
                
                #skip = rec.get('Skip')
                #if skip and skip.startswith('#'):
                #    continue
                reclst.append(rec)
            
            rowidx += 1

        resultlst.append(reclst)
        
    return resultlst
def savesheet(fn,wb,title,descriptor,data):
    ws = wb.create_sheet(title=title)
    ws.append([col[0] for col in descriptor])
    for rec in data:
        ws.append([rec.get(col[1],"") for col in descriptor])
    wb.save(fn)   
def main(args):
    fn = r'./HF_IDUCENTER-icd.xlsx'
    inputsignals = []
    wb = load_workbook(fn)
    afdxinputsignals, a825inputsignals = readExcelFile(wb,('AFDXInputSignals','A825InputSignals'))
    inputsignals.extend(afdxinputsignals)
    inputsignals.extend(a825inputsignals)
    savesheet(fn,wb,'inputsignals',signalsInputColumns,inputsignals)

if __name__ == '__main__':
   sys.exit(main(sys.argv[1:]))     
    
    