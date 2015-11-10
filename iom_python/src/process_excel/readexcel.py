from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, Font, Side, Border, PatternFill, Style
from openpyxl.reader.excel import load_workbook

filename = r'./HF_IDUCENTER-icd.xlsx'
wb = load_workbook(filename)
ws = wb.create_sheet(title="InputSignals") 
wsafdx =  wb.get_sheet_by_name("AFDXInputSignals")
#for row in wsafdx.rows:
#    for cell in row:
#        cell.index
for i in range(26):
    for j in range(1700):
        col = get_column_letter(i+1)
        ws.cell('%s%s'%(col,j+1)).value = wsafdx.cell('%s%s'%(col,j+1)).value
wb.save(filename)